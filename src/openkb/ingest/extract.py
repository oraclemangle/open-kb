"""Text extraction — turn a source document into plain text (or TSV-ish rows).

Every ingestible format lands here as one function that takes a path and
returns a string. The worker decides what happens *after* extraction
(classify, chunk, embed) — this module only answers "what text is in this
file?", plus, for PDFs, cheap per-page signal so the caller can decide
whether OCR is needed without re-opening the file.

Design notes:
  * PDF text extraction is done with PyMuPDF (``fitz``), imported lazily
    inside the function that needs it. PyMuPDF is an optional dependency
    (see ``pyproject.toml``'s ``pdf`` extra) — importing it at module load
    time would make every caller of this module require it, even ones that
    only ever touch .txt/.csv files. A missing import should fail with a
    clear, actionable message at the point of use, not a cryptic
    ModuleNotFoundError three frames deep.
  * Office XML formats (.docx/.pptx) are zip archives of XML. We walk the
    zip and pull text nodes out of the relevant XML parts by hand rather
    than pulling in a heavyweight document-object-model library, but a zip
    file from an unknown source is an attack surface (zip bombs: a small
    file that expands to gigabytes, or a compression ratio calculated to
    exhaust memory/CPU). We guard against both an absurd entry count and an
    absurd single-entry size/ratio before ever inflating anything.
  * .xlsx/.xlsm has two extraction modes. The default ("generic") walks
    every sheet with openpyxl and emits loose TSV-ish lines — fine for
    small workbooks used as reference material. "Register" mode
    (``extract_register``) is for the opposite case: a single large,
    strictly tabular sheet (a cable schedule, an I/O list, an asset
    register) where the header row must be found precisely and every data
    row must line up under it. The naive per-cell text dump that a generic
    XML/shared-strings scrape would produce is *worse than useless* for a
    register that size: with thousands of numeric-looking cells and no
    column alignment, the extracted text degenerates into a wall of
    numbers with no way to tell "port 12" from "cable run 12" from
    "rack unit 12". Aligning every row under its header, tab-separated,
    is what makes the register greppable and chunkable in a way that
    actually answers questions later.
"""
from __future__ import annotations

import datetime as _dt
import os
import re
import zipfile

OFFICE_XML_KEYWORDS = ("document", "sheet", "slide", "shared")
ZIP_MAX_ENTRIES = 5000
ZIP_MAX_XML_BYTES = 50_000_000
ZIP_MAX_RATIO = 200

IMAGE_EXTENSIONS = {".png", ".jpg", ".jpeg"}

__all__ = [
    "extract_pdf_text",
    "extract_office_xml",
    "extract_xlsx_generic",
    "extract_register",
    "extract_plain_text",
    "extract_text",
]


def _require_fitz():
    try:
        import fitz  # PyMuPDF
    except ImportError as exc:  # pragma: no cover - depends on optional install
        raise RuntimeError(
            "PyMuPDF is required for PDF handling. Install the 'pdf' extra: "
            "pip install 'open-kb[pdf]'"
        ) from exc
    return fitz


def extract_pdf_text(path: str) -> tuple[str, list[int]]:
    """Extract the text layer of every page in a PDF.

    Returns ``(text, chars_per_page)``. ``chars_per_page`` lets the caller
    (the ingest worker) decide, per document, whether the text layer is
    real content or just a thin sprinkling of title-block text over a
    scanned/graphic page — the signal that should trigger OCR. We don't
    make that decision here: extraction and the OCR-fallback policy are
    different concerns, and the worker also knows the configured
    thresholds (``ocr.*``), which this module intentionally does not.
    """
    fitz = _require_fitz()
    doc = fitz.open(path)
    try:
        parts: list[str] = []
        chars_per_page: list[int] = []
        for page in doc:
            try:
                text = (page.get_text() or "").strip()
            except Exception:
                text = ""
            parts.append(text)
            chars_per_page.append(len(text))
        return "\n\n".join(parts), chars_per_page
    finally:
        doc.close()


def _zip_guard(zf: zipfile.ZipFile) -> None:
    """Reject zip archives shaped like a zip bomb before inflating anything."""
    infos = zf.infolist()
    if len(infos) > ZIP_MAX_ENTRIES:
        raise ValueError("zip has %d entries (limit %d)" % (len(infos), ZIP_MAX_ENTRIES))
    for info in infos:
        if info.file_size > ZIP_MAX_XML_BYTES:
            raise ValueError("zip entry %r too large uncompressed (%d bytes)" % (info.filename, info.file_size))
        if info.compress_size > 0 and info.file_size / info.compress_size > ZIP_MAX_RATIO:
            raise ValueError("zip entry %r has a suspicious compression ratio" % info.filename)


def extract_office_xml(path: str) -> str:
    """Extract visible text from a .docx/.pptx by reading its XML parts directly.

    .docx and .pptx are zip archives containing XML documents (WordprocessingML
    / PresentationML). Rather than depend on a full document-object-model
    library, we open the archive, guard against zip-bomb shapes, and pull
    the ``.text`` of every XML element out of the parts that actually carry
    document content (word/document.xml, ppt/slides/slideN.xml, and the
    shared-strings-equivalent parts). This loses structure (no paragraph
    breaks, no table cells) but keeps every word — good enough for
    retrieval, which only needs the words to be present and chunkable.

    XML parser choice: we prefer ``defusedxml`` (guards against XML entity
    expansion / external entity attacks) and fall back to the stdlib
    ``xml.etree.ElementTree`` if it isn't installed. The stdlib parser is
    not hardened against malicious XML (billion-laughs, external entity
    fetches) — that's an accepted trade-off for a document you already
    trust enough to have chosen to ingest, and it's why the zip-bomb guard
    above runs unconditionally regardless of which parser is available.
    Install ``defusedxml`` if you plan to ingest documents from untrusted
    sources.
    """
    try:
        from defusedxml import ElementTree as ET  # type: ignore
    except ImportError:
        from xml.etree import ElementTree as ET  # nosec - see docstring trade-off note

    texts: list[str] = []
    with zipfile.ZipFile(path) as zf:
        _zip_guard(zf)
        for info in zf.infolist():
            name = info.filename
            if not name.endswith(".xml"):
                continue
            if not any(kw in name for kw in OFFICE_XML_KEYWORDS):
                continue
            try:
                root = ET.fromstring(zf.read(name))
            except Exception:
                continue
            for el in root.iter():
                if el.text and el.text.strip():
                    texts.append(el.text.strip())
    return "\n".join(texts)


def _cell_to_str(value) -> str:
    if value is None:
        return ""
    if isinstance(value, (_dt.datetime, _dt.date)):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, float):
        return str(int(value)) if value.is_integer() else ("%.6g" % value)
    return str(value).strip()


def extract_xlsx_generic(path: str, max_rows: int = 8000, max_cols: int = 64) -> str:
    """Extract every sheet of a workbook as loose, sheet-labelled TSV-ish text.

    Uses openpyxl in read-only/data-only mode so formulas resolve to their
    last-calculated values (not the formula text) and shared strings resolve
    to real cell content — a naive raw-XML scrape of a workbook returns
    shared-string *indices*, which are meaningless numbers without the
    lookup table alongside them. This is the right default for reference
    workbooks (a handful of small sheets); for one very large, strictly
    tabular sheet where row/column alignment must be exact and searchable
    (a register), use :func:`extract_register` instead.
    """
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    out: list[str] = []
    try:
        for ws in wb.worksheets:
            rows: list[str] = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i >= max_rows:
                    rows.append("... (rows truncated)")
                    break
                cells = [_cell_to_str(v) for v in row[:max_cols]]
                while cells and cells[-1] == "":
                    cells.pop()
                if any(cells):
                    rows.append(" | ".join(cells))
            if rows:
                out.append("## Sheet: %s (%d rows)\n%s" % (ws.title, len(rows), "\n".join(rows)))
    finally:
        wb.close()
    return "\n\n".join(out)


def _normalise_cell(value) -> str:
    """Collapse embedded newlines/whitespace in a cell so one row stays one line."""
    s = _cell_to_str(value)
    return " ".join(s.split())


def extract_register(
    path: str,
    header_keyword: str,
    sheet: str | None = None,
    rows_cap: int | None = None,
) -> tuple[list[str], list[list[str]]]:
    """Extract one large tabular sheet as a precisely header-aligned table.

    Registers (a cable schedule, a switch-port list, an asset/equipment
    master list) are exactly the case where the generic sheet dump falls
    down: thousands of rows of numbers and short codes, no natural-language
    context per cell, and a title/notes row or two before the real header.
    If those rows aren't lined up under their column names, the extracted
    text is an undifferentiated wall of numbers — nothing in it says which
    number is a port, a cable ID, or a rack position, so it cannot be
    grepped, chunked usefully, or explain itself out of context.

    This finds the header row by two independent signals — at least 3
    non-empty cells (skips one-cell title rows like "Cable Schedule 2024")
    AND a case-insensitive match for ``header_keyword`` somewhere in the
    row (skips notes/instruction rows that happen to have several cells)
    — then reads every following non-blank row and pads/truncates it to
    the header's width. Every value is passed through
    :func:`_normalise_cell` to collapse embedded line breaks, so a
    multi-line cell (wrapped notes) still counts as one table row.

    ``rows_cap`` limits how many data rows are returned, useful for a
    quick preview of a huge register before committing to embedding all
    of it. ``sheet=None`` uses the first sheet in the workbook.

    Returns ``(header, rows)``; ``header`` is ``[]`` if the keyword was
    never found (caller should treat that as "not a register, fall back
    to generic extraction").
    """
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[sheet] if sheet else wb.worksheets[0]
        header: list[str] | None = None
        rows: list[list[str]] = []
        keyword = header_keyword.lower()
        for raw_row in ws.iter_rows(values_only=True):
            cells = [_normalise_cell(v) for v in raw_row]
            if header is None:
                non_empty = sum(1 for c in cells if c)
                if non_empty >= 3 and any(keyword in c.lower() for c in cells if c):
                    header = list(cells)
                continue
            if not any(cells):
                continue
            rows.append(cells)
            if rows_cap and len(rows) >= rows_cap:
                break
    finally:
        wb.close()

    if header is None:
        return [], []

    while header and not header[-1]:
        header.pop()
    width = len(header)
    aligned = [r[:width] + [""] * (width - len(r)) for r in rows]
    return header, aligned


def register_to_tsv_rows(header: list[str], rows: list[list[str]]) -> list[str]:
    """Render a header + data rows as tab-separated lines (header first)."""
    lines = ["\t".join(header)]
    lines.extend("\t".join(r) for r in rows)
    return lines


def extract_plain_text(path: str) -> str:
    """Read a .txt/.md/.csv file directly as UTF-8, tolerating bad bytes."""
    with open(path, "r", encoding="utf-8", errors="ignore") as fh:
        return fh.read()


def extract_text(path: str, cfg: dict) -> tuple[str, str]:
    """Dispatch extraction by extension. Returns ``(text, method)``.

    ``method`` is one of ``"pdf"``, ``"office"``, ``"xlsx"``, ``"text"``,
    ``"image"`` (the last meaning "route to OCR — no text here yet") — a
    hint the worker uses when deciding the ``documents.extractor`` value
    and whether OCR fallback applies. This function never calls the OCR/VLM
    endpoints itself (see ``ingest.ocr``); for images it deliberately
    returns an empty string so the worker's OCR branch is the single place
    that decides to spend a vision-model call.
    """
    ext = os.path.splitext(path)[1].lower()
    if ext == ".pdf":
        text, _chars_per_page = extract_pdf_text(path)
        return text, "pdf"
    if ext in IMAGE_EXTENSIONS:
        return "", "image"
    if ext in (".xlsx", ".xlsm"):
        registers = cfg.get("ingest", {}).get("registers") or []
        for reg in registers:
            pattern = reg.get("glob")
            if pattern and _path_matches_glob(path, pattern):
                header, rows = extract_register(
                    path,
                    header_keyword=reg.get("header_keyword", ""),
                    sheet=reg.get("sheet"),
                )
                if header:
                    return "\n".join(register_to_tsv_rows(header, rows)), "xlsx-register"
        return extract_xlsx_generic(path), "xlsx"
    if ext in (".docx", ".pptx"):
        return extract_office_xml(path), "office"
    if ext in (".txt", ".md", ".csv"):
        return extract_plain_text(path), "text"
    return "", "unsupported"


def _path_matches_glob(path: str, pattern: str) -> bool:
    """Match an absolute/relative path against a ``**``-style glob pattern."""
    import fnmatch

    norm_path = path.replace(os.sep, "/")
    return fnmatch.fnmatch(norm_path, pattern) or fnmatch.fnmatch(os.path.basename(path), pattern)
