#!/usr/bin/env python3
"""Generate the binary artifacts of the synthetic demo corpus.

The markdown/text manuals under ``examples/corpus/`` are committed
directly (hand-written prose). This script generates the binary
artifacts that are easier to produce programmatically and keep in sync
than to hand-author byte-for-byte:

  - cable-list.xlsx        a two-sheet workbook: a real cable register
                            (title row + header row + ~80 data rows) on
                            sheet 1, and a junk notes sheet with no proper
                            header on sheet 2 (to exercise header-row
                            detection against noise).
  - datasheet.pdf           a 2-page PDF with real extractable text (title,
                            spec table, paragraphs) -- exercises normal PDF
                            text extraction.
  - graphic-only-drawing.pdf a 1-page PDF that is almost entirely vector
                            shapes with only a couple of tiny text labels,
                            so ``page.get_text()`` yields well under
                            ~40 chars -- shaped for the OCR/vision-describe
                            "text-thin" classification path.

Everything here is 100% synthetic/fictional -- see examples/corpus/README.md.

Re-runnable: this script always overwrites its own output files with the
same deterministic content, so running it twice produces byte-identical
(or near-identical, for anything that embeds a generation timestamp --
nothing here does) artifacts.
"""
from __future__ import annotations

import os

CORPUS_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "corpus")

TAGS = ["DG1", "DG2", "MSB-1", "FP-101", "CWP-01", "AHU-3"]


def make_cable_list_xlsx() -> str:
    import openpyxl
    from openpyxl.styles import Font

    path = os.path.join(CORPUS_DIR, "cable-list.xlsx")
    wb = openpyxl.Workbook()

    # --- Sheet 1: the real cable register ---
    ws = wb.active
    ws.title = "Cable Register"
    ws.append(["Facility Cable Register — DRAFT"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])  # blank spacer row
    ws.append(["Cable Number", "From", "To", "Type", "Cores", "Route"])
    for c in ws[3]:
        c.font = Font(bold=True)

    # Deterministic ~80-row register referencing the shared tags as
    # From/To endpoints, cycling through plausible cable types/routes.
    endpoints = [
        ("DG1", "MSB-1"), ("DG2", "MSB-1"), ("MSB-1", "FP-101"),
        ("MSB-1", "AHU-3"), ("MSB-1", "CWP-01"), ("FP-101", "CWP-01"),
        ("CWP-01", "AHU-3"), ("DG1", "DG2"),
    ]
    cable_types = ["4C XLPE 95mm2", "4C XLPE 50mm2", "3C+E SWA 25mm2", "Cat6 SFTP", "2C Control 1.5mm2"]
    routes = ["Cable tray CT-1 -> CT-4", "Underfloor duct D-2", "Cable tray CT-3", "Riser R-1", "Direct buried"]

    n_rows = 80
    for i in range(1, n_rows + 1):
        frm, to = endpoints[i % len(endpoints)]
        cable_no = "CAB-%03d" % i
        ctype = cable_types[i % len(cable_types)]
        cores = str(2 + (i % 4))
        route = routes[i % len(routes)]
        ws.append([cable_no, frm, to, ctype, cores, route])

    # --- Sheet 2: junk notes sheet, no proper header row ---
    junk = wb.create_sheet("Notes (ignore)")
    junk.append(["misc notes -- do not use for extraction"])
    junk.append([])
    junk.append(["check spare drum lengths with the stores office"])
    junk.append(["", "", "some stray value", ""])
    junk.append(["TBD"])

    wb.save(path)
    return path


def make_datasheet_pdf() -> str:
    import fitz  # PyMuPDF

    path = os.path.join(CORPUS_DIR, "datasheet.pdf")
    doc = fitz.open()

    # Page 1: title + spec table as text
    page = doc.new_page()
    page.insert_text((72, 90), "Aurora Power Systems APG-500", fontsize=18, fontname="helv")
    page.insert_text((72, 115), "Standby Diesel Generator — Technical Datasheet", fontsize=12)
    page.insert_text((72, 150), "Rated Output", fontsize=11, fontname="helv")
    page.insert_text((250, 150), "500 kVA / 400 kW @ 0.8 PF", fontsize=11)
    page.insert_text((72, 172), "Voltage / Frequency", fontsize=11)
    page.insert_text((250, 172), "400V 3-phase, 50 Hz", fontsize=11)
    page.insert_text((72, 194), "Prime Mover", fontsize=11)
    page.insert_text((250, 194), "6-cylinder turbocharged diesel, water-cooled", fontsize=11)
    page.insert_text((72, 216), "Starting System", fontsize=11)
    page.insert_text((250, 216), "24V DC electric start", fontsize=11)
    page.insert_text((72, 238), "Fuel Tank (base)", fontsize=11)
    page.insert_text((250, 238), "400 L integral base tank", fontsize=11)
    page.insert_text((72, 260), "Cooling", fontsize=11)
    page.insert_text((250, 260), "Radiator, engine-driven fan", fontsize=11)
    page.insert_text(
        (72, 300),
        "The APG-500 is designed for standby/emergency duty in facilities\n"
        "requiring automatic mains-failure changeover. Typical installation\n"
        "pairs two units for alternating weekly duty against a single\n"
        "facility load, each unit individually rated for the full load.",
        fontsize=10,
    )

    # Page 2: further specs + paragraph
    page2 = doc.new_page()
    page2.insert_text((72, 90), "Dimensions & Weight", fontsize=14, fontname="helv")
    page2.insert_text((72, 120), "Length", fontsize=11)
    page2.insert_text((250, 120), "3400 mm", fontsize=11)
    page2.insert_text((72, 142), "Width", fontsize=11)
    page2.insert_text((250, 142), "1200 mm", fontsize=11)
    page2.insert_text((72, 164), "Height", fontsize=11)
    page2.insert_text((250, 164), "1850 mm", fontsize=11)
    page2.insert_text((72, 186), "Dry Weight", fontsize=11)
    page2.insert_text((250, 186), "3950 kg", fontsize=11)
    page2.insert_text(
        (72, 230),
        "Warranty and Service\n\n"
        "Aurora Power Systems warrants the APG-500 for 24 months from\n"
        "commissioning date or 2000 running hours, whichever occurs first.\n"
        "Scheduled maintenance intervals are documented in the facility's\n"
        "own equipment manual for the installed unit.",
        fontsize=10,
    )

    doc.save(path)
    doc.close()
    return path


def make_graphic_only_pdf() -> str:
    import fitz  # PyMuPDF

    path = os.path.join(CORPUS_DIR, "graphic-only-drawing.pdf")
    doc = fitz.open()
    page = doc.new_page()

    # A handful of vector shapes standing in for a schematic -- no prose,
    # just boxes/lines representing equipment outlines and connections.
    shapes = page.new_shape()
    shapes.draw_rect(fitz.Rect(80, 100, 200, 180))    # DG1 outline
    shapes.draw_rect(fitz.Rect(80, 220, 200, 300))    # DG2 outline
    shapes.draw_rect(fitz.Rect(320, 160, 460, 240))   # MSB-1 outline
    shapes.draw_line(fitz.Point(200, 140), fitz.Point(320, 190))
    shapes.draw_line(fitz.Point(200, 260), fitz.Point(320, 210))
    shapes.draw_rect(fitz.Rect(540, 160, 660, 240))   # downstream board outline
    shapes.draw_line(fitz.Point(460, 200), fitz.Point(540, 200))
    shapes.finish(width=1.2, color=(0, 0, 0))
    shapes.commit()

    # Only a couple of tiny tag labels -- deliberately far under the
    # ~40-chars/page "text-thin" threshold used by the ingest pipeline.
    page.insert_text((110, 95), "DG1", fontsize=8)
    page.insert_text((110, 215), "DG2", fontsize=8)
    page.insert_text((360, 155), "MSB-1", fontsize=8)

    doc.save(path)
    doc.close()
    return path


def main() -> None:
    os.makedirs(CORPUS_DIR, exist_ok=True)
    xlsx = make_cable_list_xlsx()
    pdf1 = make_datasheet_pdf()
    pdf2 = make_graphic_only_pdf()
    print("Generated:")
    for p in (xlsx, pdf1, pdf2):
        print("  %s (%d bytes)" % (p, os.path.getsize(p)))


if __name__ == "__main__":
    main()
